from lxml import etree
import openpyxl as op
import time


start = time.time()


def convert():
    filename = 'rtk.xlsx'
    wb = op.load_workbook(filename)
    sheet = wb.worksheets[0]
    max_row = sheet.max_row
    for rows in sheet.iter_rows(min_row=1, max_col=45, max_row=128):
        for row in rows:
            if str(row.value).startswith('Итого'):
                print(row.value)
    end = time.time() - start
    print(end)

if __name__ == '__main__':
    convert()
