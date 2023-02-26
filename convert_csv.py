import csv
import time
import os
import openpyxl


start = time.time()
departments_numbers = {}
departments_files = os.listdir('departments')
# print(departments_files)
wb = openpyxl.Workbook()
os.chdir('departments')
# Делаем словарь из филиалов и номеров телефонов
for df in departments_files:
    with open(df, 'r') as file:
        # print(file.name)
        phone_numbers = file.read().splitlines()
        departments_numbers[file.name] = phone_numbers
        sheet = wb.create_sheet(file.name.split('.')[0])
# print(departments_numbers)


def convert():
    os.chdir('..')
    filename = 'rtk.csv'
    result = []
    # перебираем файл и собираем звонки в список
    with open(filename, 'r', encoding='utf-8') as csv_file:
        reader = csv.reader(csv_file)
        for row in reader:
            if str(row).find('Итого т.') != -1:
                a = row[0].split(';')
                if len(row) > 1:
                    # print(row[])
                    a.append(row[1].split(';')[0])
                if a[11] == '':
                    a[11] = '0'
                result.append([a[0].replace('Итого т.', ''), int(a[10]) + int(a[11])/100])
    # print(len(result))
    # print(result)
    csv_file.close()
    # перебираем звонки и распределяем по филиам
    # заполняем таблицу Excel по филиалам
    to_remove = []
    for ind, sums in enumerate(result):
        for dep, nums in departments_numbers.items():
            # print(f'Sums - {sums}, Nums - {nums}')
            if sums[0] in nums:
                # print(f'{dep} {sums[0]}: {sums[1]}')
                sheet = wb[dep.split('.')[0]]
                num_cell = sheet.cell(row=sheet.max_row + 1, column=1)
                num_cell.value = sums[0]
                money_cell = sheet.cell(row=sheet.max_row, column=2)
                money_cell.value = sums[1]
                to_remove.append(ind)
    print(len(to_remove))
    result[:] = [x for i,x in enumerate(result) if i not in to_remove]
    # выводим отсев
    sheet = wb['Sheet']
    sheet.title = 'Отсев'
    for o in result:
        ots_cell = sheet.cell(row=sheet.max_row+1, column=1)
        ots_cell.value = o[0]
        ots_cell = sheet.cell(row=sheet.max_row, column=2)
        ots_cell.value = o[1]
    wb.save('output.xlsx')
    print(len(result))


if __name__ == '__main__':
    convert()
    end = time.time() - start
    print(end)
